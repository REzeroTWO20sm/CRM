import openpyxl

from openpyxl import load_workbook

wb = load_workbook('data_crm.xlsx')
sheet_ranges = wb['Лист1']

dish_id = sheet_ranges['A']
dish_name = sheet_ranges['B']
dish_description = sheet_ranges['C']
dish_price = sheet_ranges['D']
dish_category_id_2 = sheet_ranges['E']

dish_category_id_1 = sheet_ranges['G']
dish_category_name = sheet_ranges['H']

# -------------------------------------------------------------------------------------------
# for category_dish 'insert into category_dish (name) values ('
#
# sql_command = 'insert into category_dish (name) values ('
#
# for row in sheet_ranges.iter_rows(min_row=3, max_row=23, min_col=7, max_col=8):
#     a=0
#     for cell in row:
#         a+=1
#         if a==1:
#             sql_command += ""
#         elif a==2:
#             sql_command += "'"+str(cell.value)+"'"+");"
#             print(sql_command,"\n")
#     sql_command = 'insert into category_dish (name) values ('
# -------------------------------------------------------------------------------------------

# -------------------------------------------------------------------------------------------
# for dish 'insert into dish (name,description,category_dish_id,recipe) values ('
#
# sql_command = 'insert into dish (name,description,category_dish_id,recipe) values ('

# for row in sheet_ranges.iter_rows(min_row=3, max_row=98, min_col=1, max_col=5):
#     a=0
#     for cell in row:
#         a+=1
#         if a==1:
#             sql_command += ""
#         elif a==2 or a==3:
#             sql_command += "'"+str(cell.value)+"'"+","
#         elif a==4:
#             sql_command += str(cell.value) + ","
#         elif a==5:
#             sql_command += "'"+str(cell.value)+"'"+");"
#             print(sql_command,"\n")
#     sql_command = 'insert into dish (name,description,category_dish_id,recipe) values ('
# -------------------------------------------------------------------------------------------


sql_command = 'insert into dish (name,description,category_dish_id,recipe) values ('

for row in sheet_ranges.iter_rows(min_row=3, max_row=98, min_col=1, max_col=5):
    a=0
    for cell in row:
        a+=1
        if a==1:
            sql_command += ""
        elif a==2 or a==3:
            sql_command += "'"+str(cell.value)+"'"+","
        elif a==4:
            sql_command += str(cell.value) + ","
        elif a==5:
            sql_command += "'"+str(cell.value)+"'"+");"
            print(sql_command,"\n")
    sql_command = 'insert into dish (name,description,category_dish_id,recipe) values ('